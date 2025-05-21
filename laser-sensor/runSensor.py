import sys
import time
import threading
from queue import Queue, Empty

from PyQt5 import QtWidgets, QtCore
import pyqtgraph as pg
import serial
import serial.tools.list_ports

import os
import csv
from datetime import datetime
import pandas as pd
from PyQt5.QtWidgets import QFileDialog, QMessageBox
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl import Workbook
from openpyxl.styles import PatternFill  # Changed from openpyxl.drawing.fill
from datetime import datetime
from PyQt5.QtWidgets import QFileDialog, QMessageBox
import numpy as np
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference, ScatterChart, Series
from openpyxl import Workbook
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.layout import Layout, ManualLayout

class SerialThread(threading.Thread):
    """
    Thread that reads COM3 at 9600 baud and enqueues ES and EN values.
    """
    def __init__(self, port='COM4', baud=9600):
        super().__init__(daemon=True)
        self.port = port
        self.baud = baud
        self.is_collecting = False
        self._stop = False
        self.queue = Queue()
        self.markers = ['EN', 'ES', 'NE', 'NW', 'WN', 'SE', 'SW', 'WS']
    def run(self):
        try:
            ser = serial.Serial(self.port, self.baud, timeout=1)
            print(f"[SerialThread] Opened {self.port} at {self.baud} baud.")
        except Exception as e:
            print(f"[SerialThread] ERROR opening {self.port}: {e}")
            return

        while not self._stop:
            if self.is_collecting and ser.in_waiting:
                line = ser.readline().decode('utf-8', errors='ignore').strip()
                if line:
                    # parse out ALL SENSOR values
                    parts = line.split('|')
                    ts = time.time()
                    for chunk in parts:
                        for marker in self.markers:
                            if chunk.startswith(marker):
                                try:
                                    val = float(chunk[len(marker):])
                                    self.queue.put((marker, ts, val))
                                except ValueError:
                                    pass
            else:
                time.sleep(0.05)

        ser.close()
        print("[SerialThread] Serial connection closed.")

    def stop(self):
        self._stop = True


class MainWindow(QtWidgets.QMainWindow):
    """
    GUI with two real-time plots for ES and EN displacement.
    """
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Laser Sensors Plots")
        self.resize(1000, 900)

        cw = QtWidgets.QWidget()
        self.setCentralWidget(cw)
        layout = QtWidgets.QVBoxLayout(cw)
        
        self.markers = ['EN', 'ES', 'NE', 'NW', 'WN', 'SE', 'SW', 'WS']
        self.part_data = {marker: [] for marker in self.markers}
        self.plots = {}
        self.curves = {}

        grid = QtWidgets.QGridLayout()
        for idx, marker in enumerate(self.markers):
            plot = pg.PlotWidget(title=f"{marker}")
            plot.setLabel('bottom', 'Time (s)')
            plot.setLabel('left',   'Displacement (in)')
            curve = plot.plot([], [], pen=pg.mkPen(width=2))
            self.plots[marker] = plot
            self.curves[marker] = curve
            grid.addWidget(plot, idx // 2, idx % 2)
        layout.addLayout(grid)

        # # ES plot
        # self.plot_es = pg.PlotWidget(title="ES Displacement")
        # self.plot_es.setLabel('bottom', 'Time Elapsed (s)')
        # self.plot_es.setLabel('left',   'ES Value (in)')
        # layout.addWidget(self.plot_es)
        # self.es_curve = self.plot_es.plot([], [], pen=pg.mkPen('y', width=2))
        # self.es_data = []

        # # EN plot
        # self.plot_en = pg.PlotWidget(title="EN Displacement")
        # self.plot_en.setLabel('bottom', 'Time Elapsed (s)')
        # self.plot_en.setLabel('left',   'EN Value (in)')
        # layout.addWidget(self.plot_en)
        # self.en_curve = self.plot_en.plot([], [], pen=pg.mkPen('r', width=2))
        # self.en_data = []

        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()
        self.start_btn = QtWidgets.QPushButton("Start")
        self.stop_btn  = QtWidgets.QPushButton("Stop")
        self.save_btn  = QtWidgets.QPushButton("Save")  
        self.reset_btn = QtWidgets.QPushButton("Reset")  
        self.stop_btn.setEnabled(False)
        self.save_btn.setEnabled(False)  
        btn_layout.addWidget(self.start_btn)
        btn_layout.addWidget(self.stop_btn)
        btn_layout.addWidget(self.save_btn)  
        btn_layout.addWidget(self.reset_btn)  
        layout.addLayout(btn_layout)

        # Status
        self.status = QtWidgets.QLabel("Status: Idle")
        layout.addWidget(self.status)

        # Serial thread
        self.serial_thread = SerialThread(port='COM4', baud=9600) #this is HARDCODED
        self.serial_thread.start()

        # Timer to update both plots
        self.timer = QtCore.QTimer(self)
        self.timer.setInterval(100)
        self.timer.timeout.connect(self.update_plots)
        self.timer.start()

        # Signals
        self.start_btn.clicked.connect(self.start_recording)
        self.stop_btn.clicked.connect(self.stop_recording)
        self.save_btn.clicked.connect(self.save_to_excel)
        self.reset_btn.clicked.connect(self.reset_data)  

        self.first_start = True  

    def start_recording(self):
        
        if self.first_start:
            self.start_time = time.time()
            for marker in self.markers:
                self.part_data[marker].clear()
                self.curves[marker].setData([], [])
            self.status.setText("Status: Data initialized and recording")
            self.first_start = False
        else:
            self.status.setText("Status: Recording")
        self.serial_thread.is_collecting = True
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.save_btn.setEnabled(True)

    def stop_recording(self):
        self.serial_thread.is_collecting = False
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status.setText("Status: Stopped")

    def reset_data(self):
        for marker in self.markers:
            self.part_data[marker].clear()
            self.curves[marker].setData([], [])
        self.status.setText("Status: Data Reset")
        self.first_start = True
        self.save_btn.setEnabled(False)  
    
    

    def save_to_excel(self):
        import numpy as np

        def extract_linear_dataset(x, y,
                                slope_range,
                                window_size=25,
                                step=0.1):
            
            x = np.asarray(x)
            y = np.asarray(y)
            N = len(x)
            if N < window_size:
                raise ValueError("Data shorter than window size")

            # --- STEP A: detect segment (as before) ---
            valid = []
            smin, smax = slope_range
            for i in range(N - window_size + 1):
                xs = x[i:i+window_size]
                ys = y[i:i+window_size]
                m, _ = np.polyfit(xs, ys, 1)
                if smin <= m <= smax:
                    valid.append(i)
            if not valid:
                raise ValueError("No segment found in slope range")

            # group runs of consecutive window‐starts
            runs, run = [], [valid[0]]
            for idx in valid[1:]:
                if idx == run[-1] + 1:
                    run.append(idx)
                else:
                    runs.append(run); run = [idx]
            runs.append(run)
            best = max(runs, key=len)
            first_win, last_win = best[0], best[-1]

            # convert to true data indices (drop window edges)
            start_idx = first_win + window_size - 1
            end_idx   = last_win

            # --- STEP B: fit one line to that segment ---
            seg_x = x[start_idx:end_idx+1]
            seg_y = y[start_idx:end_idx+1]
            slope, intercept = np.polyfit(seg_x, seg_y, 1)

            # --- STEP C: build new (x,y) along that line ---
            x0, x1 = x[start_idx], x[end_idx]
            # include x0, then steps of 'step' up to <= x1
            new_x = np.arange(x0, x1 + step/2, step)
            new_y = slope * new_x + intercept

            new_data = list(zip(new_x.tolist(), new_y.tolist()))
            return new_data, start_idx, end_idx, slope, intercept

    
        if not any(self.part_data[m] for m in self.markers):
            QMessageBox.warning(self, "Save Error", "No data to save.")
            return

        # Create filename with timestamp
        ts = datetime.now().strftime("%Y_%m_%d_%H%M%S")
        default_name = f"sensor_data_{ts}.xlsx"
        out_fn, _ = QFileDialog.getSaveFileName(self, "Save Excel File", default_name, "Excel Files (*.xlsx)")
        if not out_fn:
            return

        # Create workbook
        wb = openpyxl.Workbook()

        # Process each sensor's data
        for m in self.markers:
            data = self.part_data[m]
            if not data:
                continue

            # Create worksheet for each sensor
            if m == self.markers[0]:
                ws = wb.active
                ws.title = m
            else:
                ws = wb.create_sheet(title=m)

            # Add headers
            ws['A1'] = 'Time (s)'
            ws['B1'] = 'Displacement (in)'
            ws['C1'] = 'Fitted Time'
            ws['D1'] = 'Fitted Displacement'
            ws[f'E{1}'] = "Fitted Slope (in/s)"
            ws[f'E{4}'] = "Speed (in/min)"
            
            # Adjust column widths to fit content
            for column in ['A', 'B', 'D', 'E', 'F']:
                max_length = 0
                column_letter = column
                for cell in ws[column]:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            # Write data
            numData = len(data)
            for row, (time_val, disp_val) in enumerate(data, start=2):
                ws[f'A{row}'] = time_val
                ws[f'B{row}'] = disp_val

            # Create scatter chart
            chart = ScatterChart()
            chart.title = f"{m} Displacement vs Time"
            chart.height = 15
            chart.width = 25
            chart.x_axis.title = "Time (s)"
            chart.y_axis.title = "Displacement (in)"
            
            chart.x_axis.delete = False
            chart.y_axis.delete = False

            chart.x_axis.majorGridlines = None

            chart.y_axis.scaling.min = -10     
            chart.y_axis.scaling.max = 70  

            if chart.plot_area.layout is None:
                chart.plot_area.layout = Layout()

            chart.plot_area.layout = Layout(
                manualLayout=ManualLayout(
                    x=0.15,   # fraction from left edge (0.0–1.0)
                    y=0.10,   # fraction from top edge  (0.0–1.0)
                    w=0.70,   # width  as fraction of chart area
                    h=0.75    # height as fraction of chart area
                )
)
            # Create data references
            xvals = Reference(ws, min_col=1, min_row=2, max_row=numData+1)
            yvals = Reference(ws, min_col=2, min_row=2, max_row=numData+1)
            
            slope_range = (-100, 0)  #min , max

            # Convert References to numpy arrays first
            x_data = [ws[f'A{i}'].value for i in range(2, numData+2)]
            y_data = [ws[f'B{i}'].value for i in range(2, numData+2)]
            
            

            # Create series with explicit x and y values
            series = Series(values=yvals, xvalues=xvals, title="")
            chart.series.append(series)
            
            


            # Style the markers - set consistent color
            series.marker.symbol = "circle"
            series.marker.size = 2
            series.marker.graphicalProperties.solidFill = "0000FF"  # Blue color
            series.marker.graphicalProperties.line.solidFill = "0000FF"  # Blue outline
            series.graphicalProperties.line.noFill = True  # Remove connecting lines
            series.smooth = None  # Ensure no line smoothing
            
            try:
                # Then call extract_linear_dataset with the arrays
                fit_data, start_idx, end_idx, slope, intercept = extract_linear_dataset(
                    np.array(x_data), 
                    np.array(y_data), 
                    slope_range
                )
                
                # Write fitted line data to worksheet
                for i, (x, y) in enumerate(fit_data, start=2):
                    ws[f'C{i}'] = x
                    ws[f'D{i}'] = y
                
                # Create references for the fitted line data
                fit_xvals = Reference(ws, min_col=3, min_row=2, max_row=len(fit_data)+1)
                fit_yvals = Reference(ws, min_col=4, min_row=2, max_row=len(fit_data)+1)
                
                
                # Add markers for points used in linear fit
                series3 = Series(
                    values=Reference(ws, min_col=2, min_row=start_idx+2, max_row=end_idx+2),
                    xvalues=Reference(ws, min_col=1, min_row=start_idx+2, max_row=end_idx+2),
                    title="Points Used in Fit"
                )
                chart.series.append(series3)
                
                # Style the fitted points
                series3.marker.symbol = "diamond"
                series3.marker.size = 7
                series3.marker.graphicalProperties.solidFill = "FF0000"  # Red color
                series3.graphicalProperties.line.noFill = True
                
                
                print(f"Slope: {slope}")
                ws[f'E{2}'] = slope
                ws[f'E{5}'] = slope * 60
                
            except Exception as e:
                print(f"Error fitting line: {e}")
                pass


            # Hide legend
            chart.legend = None
            
            # Add chart to worksheet
            ws.add_chart(chart, "G2")
            

        # Save workbook
        wb.save(out_fn)
        QMessageBox.information(self, "Saved", f"Data and charts saved to:\n{out_fn}")


    def update_plots(self):
        updated = set() #weird
        while True:
            try:
                sensor, ts, val = self.serial_thread.queue.get_nowait()
                t_rel = ts - self.start_time
                if sensor in self.part_data:
                    self.part_data[sensor].append((t_rel, val))
                    updated.add(sensor)
            except Empty:
                break

        for marker in updated:
            if self.part_data[marker]:
                x, y = zip(*self.part_data[marker])
                self.curves[marker].setData(x, y)

    def closeEvent(self, event):
        self.serial_thread.stop()
        super().closeEvent(event)


def main():
    """
    Run the application:
      pip install pyqt5 pyqtgraph pyserial
    """
    app = QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()


